"""
Generador de Master Cliente Excel.
Copia la plantilla según tipo de crédito y llena DATOS CUALITATIVOS y ESTADOS FINANCIEROS.
"""
import shutil
import tempfile
from pathlib import Path

from openpyxl import load_workbook

# Mapeo: creditType -> nombre del archivo plantilla
CREDIT_TYPE_TO_TEMPLATE = {
    "adquisicion_activos": "ByCapex.xlsx",
    "proyectos_inversion": "ByCrecimiento.xlsx",
    "capital_trabajo": "ByCT.xlsx",
}

# Hojas
SHEET_DATOS = "DATOS CUALITATIVOS"
SHEET_ESTADOS_FINANCIEROS = "ESTADOS FINANCIEROS"

# Celdas E5-E8: Score Buro, Experiencia, Formalidad, ESG
CELLS_MAPPING = {
    "E5": "creditScore",      # Score en Buro Crediticio
    "E6": "experienceYears", # Experiencia en el Giro
    "E7": "formalidad",      # Formalidad Financiera
    "E8": "esgScore",       # Metricas ESG
}

# ESTADOS FINANCIEROS: año 1->col D, año 2->col F, año 3->col H
YEAR_TO_COLUMN = ["D", "F", "H"]

# incomeStatement: (row, key)
INCOME_STATEMENT_CELLS = [
    (6, "ventas"),
    (7, "costos_de_venta"),
    (9, "gastos_de_operacion"),
    (11, "gastos_financieros"),
    (12, "otros_productos"),
    (13, "otros_gastos"),
    (15, "impuestos"),
]

# balanceSheet: (row, key)
BALANCE_SHEET_CELLS = [
    (21, "activo_circulante"),
    (22, "efectivo_y_equivalentes"),
    (23, "inventarios"),
    (24, "clientes"),
    (25, "deudores_diversos"),
    (26, "activo_fijo"),
    (27, "terreno_y_edificios"),
    (28, "maquinaria_y_equipo"),
    (29, "equipo_de_transporte"),
    (30, "depreciacion_acumulada"),
    (31, "otros_activos"),
    (33, "activo_total"),
    (35, "pasivo_circulante"),
    (36, "proveedores"),
    (37, "acreedores_diversos"),
    (38, "docs_x_pagar_cp"),
    (39, "pasivo_largo_plazo"),
    (40, "docs_x_pagar_lp"),
    (41, "otros_pasivos"),
    (42, "pasivo_total"),
    (43, "capital_social"),
    (44, "ut_ejercicios_anteriores"),
    (45, "capital_contable"),
]


def get_templates_dir() -> Path:
    return Path(__file__).parent / "templates"


def _fill_estados_financieros(wb, financial_statements: dict) -> None:
    """Rellena la hoja ESTADOS FINANCIEROS con incomeStatement y balanceSheet."""
    print(f"[DEBUG] _fill_estados_financieros: financial_statements={'SÍ' if financial_statements else 'NO'}")
    if not financial_statements:
        print("[DEBUG] financial_statements vacío, no se rellena nada")
        return

    print(f"[DEBUG] Hojas en el Excel: {wb.sheetnames}")
    sheet = None
    candidates = [SHEET_ESTADOS_FINANCIEROS, "Estados Financieros"]
    for name in candidates:
        if name in wb.sheetnames:
            sheet = wb[name]
            print(f"[DEBUG] Hoja encontrada: '{name}'")
            break
    if not sheet:
        # Fallback: buscar por coincidencia parcial (ej. "Estados" o "FINANCIEROS")
        for sname in wb.sheetnames:
            if "estado" in sname.lower() or "financiero" in sname.lower():
                sheet = wb[sname]
                print(f"[DEBUG] Hoja encontrada por coincidencia: '{sname}'")
                break
    if not sheet:
        print(f"[DEBUG] NO se encontró hoja. Nombres exactos: {wb.sheetnames}")
        return

    # Ordenar años para asignar columna: 1er año->D, 2do->F, 3er->H
    years_sorted = sorted(financial_statements.keys())
    print(f"[DEBUG] Años a rellenar: {years_sorted}")

    cells_written = 0
    for year_idx, year in enumerate(years_sorted):
        if year_idx >= len(YEAR_TO_COLUMN):
            break
        col = YEAR_TO_COLUMN[year_idx]
        data = financial_statements[year]
        if not isinstance(data, dict):
            print(f"[DEBUG] Año {year}: data no es dict, skip")
            continue

        inc = data.get("incomeStatement")
        if isinstance(inc, dict):
            for row, key in INCOME_STATEMENT_CELLS:
                val = inc.get(key)
                if val is not None:
                    sheet[f"{col}{row}"] = val
                    cells_written += 1
            print(f"[DEBUG] Año {year} col {col}: incomeStatement escrito ({len([k for k in INCOME_STATEMENT_CELLS if inc.get(k[1]) is not None])} celdas)")
        else:
            print(f"[DEBUG] Año {year}: incomeStatement no es dict (es {type(inc).__name__})")

        bal = data.get("balanceSheet")
        if isinstance(bal, dict):
            for row, key in BALANCE_SHEET_CELLS:
                val = bal.get(key)
                if val is not None:
                    sheet[f"{col}{row}"] = val
                    cells_written += 1
            print(f"[DEBUG] Año {year} col {col}: balanceSheet escrito ({len([k for k in BALANCE_SHEET_CELLS if bal.get(k[1]) is not None])} celdas)")
        else:
            print(f"[DEBUG] Año {year}: balanceSheet no es dict (es {type(bal).__name__})")

    print(f"[DEBUG] Total celdas escritas en ESTADOS FINANCIEROS: {cells_written}")


def generate_master(payload: dict) -> bytes:
    """
    Genera master_cliente.xlsx a partir del payload.
    - Selecciona plantilla según creditType
    - Copia (no modifica original)
    - Llena E5-E8 en hoja DATOS CUALITATIVOS
    - Retorna bytes del archivo generado
    """
    credit_type = payload.get("creditType")
    if not credit_type or credit_type not in CREDIT_TYPE_TO_TEMPLATE:
        raise ValueError(
            f"creditType inválido: {credit_type}. "
            f"Debe ser uno de: {list(CREDIT_TYPE_TO_TEMPLATE.keys())}"
        )

    template_name = CREDIT_TYPE_TO_TEMPLATE[credit_type]
    templates_dir = get_templates_dir()
    template_path = templates_dir / template_name

    if not template_path.exists():
        raise FileNotFoundError(f"Plantilla no encontrada: {template_path}")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        shutil.copy2(template_path, tmp.name)
        copy_path = tmp.name

    try:
        wb = load_workbook(copy_path)
        # Hoja DATOS CUALITATIVOS
        if SHEET_DATOS in wb.sheetnames:
            sheet = wb[SHEET_DATOS]
        else:
            sheet = wb.active

        for cell_ref, key in CELLS_MAPPING.items():
            value = payload.get(key)
            if value is not None:
                sheet[cell_ref] = value

        # Hoja ESTADOS FINANCIEROS
        _fill_estados_financieros(wb, payload.get("financialStatements"))

        output_path = Path(copy_path).with_name("master_cliente.xlsx")
        wb.save(output_path)
        wb.close()

        return output_path.read_bytes()
    finally:
        Path(copy_path).unlink(missing_ok=True)
